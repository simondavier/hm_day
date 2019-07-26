#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import *
import quantum780_operation
import terminalcolor as tcolor

#TESTPARAS = ['HRES', 'VRES', 'VRAT', 'VIC', 'AR', 'SCAN', 'HTOT',
#                 'HSPD', 'HSPW', 'HSPP', 'VTOT', 'VSPD', 'VSPW', 'VSPP']

class SwitchConfigOperation(object):

    def __init__(self, filename, sheetNum):
        self.filename = filename
        self.sheetNum = sheetNum
        self.wb = load_workbook(filename)
        self.wb.active
        self.mySheet = self.get_sheetNames()[sheetNum]
        self.TESTPARAS= ['HRES', 'VRES', 'VRAT', 'VIC', 'AR', 'SCAN', 'HTOT',
                 'HSPD', 'HSPW', 'HSPP', 'VTOT', 'VSPD', 'VSPW', 'VSPP']

    def saveModify(self, filename):
        self.wb.save(filename)

    def get_sheetNames(self):
        """
        get all sheetnames
        :return: a list of all sheetnames
        """
        return self.wb.sheetnames

    def get_sheet(self, sheetName):
        """
        get the sheet according to the sheetname
        :param sheetName:
        :return: a sheet obj
        """
        return self.wb[sheetName]

    def getCellValue(self, row, col):
        sheet = self.get_sheet(self.mySheet)
        return sheet.cell(row,col).value

    def setCellValue(self, row, col, newv):
        sheet = self.get_sheet(self.mySheet)
        sheet.cell(row, col, newv)

    def getRowsLenth(self):
        sheet = self.get_sheet(self.mySheet)
        return sheet.max_row

    def getColumnsLenth(self):
        sheet = self.get_sheet(self.mySheet)
        return sheet.max_column

    # def getSupportTimingCode(self, column):
    #     """
    #     return a timing code generator if mark 'x' in the column;
    #     :param column:
    #     :return: a generator
    #     """
    #     for i in range(1,self.getRowsLenth()+1):
    #         #print (self.getCellValue(i+1, 37))
    #         if 'x' == self.getCellValue(i, column):
    #             yield self.getCellValue(i,3)

    def getSupportTimingList(self, column):
        """
        Return a list of all support input timing
        :param cloumn:
        :return: result[]
        """
        result = []
        for i in range(1, self.getRowsLenth()+1):
            if 'x' == self.getCellValue(i, column):
                result.append(self.getCellValue(i,1))
        return result

    def getSupportTimingCode(self, column):
        """
        Return a list of all support input timing
        :param column:
        :return:
        """
        result = []
        for i in range(1, self.getRowsLenth()+1):
            if 'x' == self.getCellValue(i, column):
                result.append(self.getCellValue(i,3))
        return  result

    def getTimingExpect(self, qdcode):
        """
        Get all input timing expected data;
        :param qdcode:
        :return: A dic of input expected ;
        TESTPARAS = ['HRES', 'VRES', 'VRAT', 'VIC', 'HVIC', 'AR', 'SCAN', 'HTOT',
                 'HSPD', 'HSPW', 'HSPP', 'VTOT', 'VSPD', 'VSPW', 'VSPP']
        """
        result = {}
        for i in range(1, self.getRowsLenth() + 1):
            if qdcode == self.getCellValue(i, 3):
                for item in self.TESTPARAS:
                    if item == 'HRES':
                        result[item] = str(self.getCellValue(i, 16))
                    elif item == 'VRES':
                        result[item] = str(self.getCellValue(i, 23))
                    elif item == 'VRAT':
                        result[item] = ("%.2f" % float(self.getCellValue(i, 21)))
                    elif item == 'VIC':
                        result[item] = str(self.getCellValue(i, 12))
                    # elif item == 'HVIC':
                    #     result[item] = str(self.getCellValue(i, 13))
                    elif item == 'AR':
                        result[item] = self.getCellValue(i, 14)
                    elif item == 'SCAN':
                        if 'prog' == self.getCellValue(i, 20):
                            result[item] = '1'
                        elif 'int' == self.getCellValue(i, 20):
                            result[item] = '2'
                        else:
                            result[item] = self.getCellValue(i, 20)
                    elif item == 'HTOT':
                        result[item] = str(self.getCellValue(i, 15))
                    elif item == 'HSPD':
                        result[item] = str(self.getCellValue(i, 17))
                    elif item == 'HSPW':
                        result[item] = str(self.getCellValue(i, 18))
                    elif item == 'HSPP':
                        if '+' == self.getCellValue(i, 19):
                            result[item] = '1'
                        elif '-' == self.getCellValue(i, 19):
                            result[item] = '0'
                        else: result[item] = self.getCellValue(i, 19)
                    elif item == 'VTOT':
                        result[item] = str(self.getCellValue(i, 22))
                    elif item == 'VSPD':
                        result[item] = str(self.getCellValue(i, 24))
                    elif item == 'VSPW':
                        result[item] = str(self.getCellValue(i, 25))
                    elif item == 'VSPP':
                        if '+' == self.getCellValue(i, 26):
                            result[item] = '1'
                        elif '-' == self.getCellValue(i, 26):
                            result[item] = '0'
                        else: result[item] = self.getCellValue(i, 26)
                    else:
                        raise("Unknow paras was found！")
        return result

    def getTimingDectec(self, opt):
        qd = quantum780_operation.Quantum780Operation()
        """
        To get Timing paras from 
        :param opt:
        0 - generator
        1 - analyzer 
        :return: a dic of timing paras.
        """
        if '0' == opt:
            return qd.generator_timing_dump()
        if '1' == opt:
            return qd.alyz_timing_dump()

    def getEdid(self, edid, opt):
        """
        return the edid block according to the qdcode;
        :param qdcode:
        :param opt:
        0 - block0
        1 - block1
        :return:
        """
        for i in range(1, self.getRowsLenth()):
            if edid == self.getCellValue(i, 1):
                if opt == '0':
                    return self.getCellValue(i,2)
                elif opt == '1':
                    return self.getCellValue(i,3)
        else:
            raise ("No edid was found in EDID data!")

    def qdcode2edid(self,qdcode):
        for i in range(1, self.getRowsLenth()):
            if qdcode == self.getCellValue(i, 3):
                return self.getCellValue(i,4)
            else:
                continue
        else:
            raise ("Can not found the qdcode in Resolution Data!")


    def compare_result(self, expected, detected):
        final='PASS'
        for item in expected.keys():
            print("EXPECTED %s is %s and DETECTED %s is %s" % (item, expected[item], item, detected[item]))
            if item == 'VRAT':
                if abs(float(expected['VRAT']) - float(detected['VRAT'])) <= 1:
                    tcolor.cprint('PASS','GREEN')
                else:
                    tcolor.cprint('FAIL','RED')
                    final='FAIL'
                continue
            if expected[item] == detected[item]:
                tcolor.cprint('PASS', 'GREEN')
            else:
                tcolor.cprint('FAIL', 'RED')
                final = 'FAIL'
        return final

    def load_config(self):
        config={}
        for i in range(1, self.getRowsLenth()+1):
            if(self.getCellValue(i,1) and self.getCellValue(i,2)!= None):
                config[self.getCellValue(i,1)]=str(self.getCellValue(i,2))
                continue
            else:
                raise("There is empty value config in file")
        return config


if __name__ == '__main__':
    filename = "D:\\TestStandDemo\\VideoSwtich\\ConfigResolutionData.xlsx"
    edid = "D:\\TestStandDemo\\VideoSwtich\\QdEDID.xlsx"
    sco = SwitchConfigOperation(filename,0)
    #edidobj = SwitchConfigOperation(edid)
    # sheetName = sco.get_sheetNames()
    # print(sheetName)
    # print(sco.getRowsLenth())
    # print(sco.getColumnsLenth())
    # dic = sco.load_config()
    # print(dic)
    # print(dic['OutputPortType'])
    #print(sco.getCellValue(1,1))
    #print(sco.getRowsLenth())
    #print(sco.getColumnsLenth())
    #ge = sco.getSupportTiming(37)
    #ge = sco.getSupportTiming(39)
    #for code in ge:
    #    print (code)
    #timing = sco.getSupportTimingList(37)
    codes = sco.getSupportTimingList(37)
    for code in codes:
        print(code)
        #print(sco.getTimingExpect(code))
    #print(sco.getEdid('1080p60','0'))
    #print(sco.getEdid('1024x576','1'))
    # qdcode = '1080p60'
    # edidc = sco.qdcode2edid(qdcode)
    # print(edidc)
    # print(edidobj.getEdid(edidc,'0'))






